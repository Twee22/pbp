from openpyxl import Workbook
from openpyxl.worksheet.properties import WorksheetProperties
from openpyxl.drawing.image import Image

from src.config_files.config import disclaimer_1, disclaimer_2, color_1, color_2, color_3, company_name, company_email

from src.output_functions.output_helper_functions import create_team_profile

def output_profile_to_xlsx(profiles, team_name, year_for_roster, year_for_report, how_many_pa_to_appear):

    doc_name = "data/" + team_name + "/" + year_for_roster + "/" + team_name + "_batting_" + year_for_report + "_report.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Team Profile"
    team_profile = create_team_profile(profiles)

    ws["A1"] = team_name.replace("_", " ").title()

    check = create_player_profile(ws, team_profile)

    for profile in profiles:
        if profile["PA"] >= how_many_pa_to_appear:
            ws = wb.create_sheet(title=profile["name"] + "_" + profile["year"])
            # Line 1: Name
            ws["A1"] = "#{}".format(profile["number"])
            ws["C1"] = profile["f_name"]
            ws["D1"] = profile["name"]
            ws["E1"] = "Year"
            ws["F1"] = profile["year"]
            # Line 2: Height, Weight, Throws, Bats, Year
            ws["A2"] = "Class:"
            ws["B2"] = profile["class"]
            ws["C2"] = "Position:"
            ws["D2"] = profile["position"]
            ws["E2"] = "B/T:"
            ws["F2"] = profile["b_t"]
            ws["G1"] = "Height:"
            ws["H1"] = profile["height"]
            ws["G2"] = "Weight:"
            ws["H2"] = profile["weight"]

            check = create_player_profile(ws, profile)

    ws = wb.create_sheet(title="Team Leaders")
    check = create_team_leaders_page(ws, profiles)
    ws = wb.create_sheet(title="Legal Disclaimer")
    check = create_legal_page(ws)

    wb.save(filename=doc_name)
    return check

def create_player_profile(ws, profile):
    # Line 3: AVG, OBP, SLG, OPS
    ws["A3"] = "AVG"
    ws["B3"] = "{:.3f}".format(profile["AVG"])
    ws["C3"] = "OBP"
    ws["D3"] = "{:.3f}".format(profile["OBP"])
    ws["E3"] = "SLG"
    ws["F3"] = "{:.3f}".format(profile["SLG"])
    ws["A4"] = "OPS"
    ws["B4"] = "{:.3f}".format(profile["OPS"])
    ws["C4"] = "BABIP"
    ws["D4"] = "{:.3f}".format(profile["BABIP"])
    # Line 4: PA, AB, H, BB, HBP
    ws["A5"] = "PA"
    ws["B5"] = profile["PA"]
    ws["C5"] = "AB"
    ws["D5"] = profile["AB"]
    ws["E5"] = "Hits"
    ws["F5"] = profile["Hits"]
    ws["A6"] = "Walks"
    ws["B6"] = profile["BB"]
    ws["C6"] = "HBP"
    ws["D6"] = profile["HBP"]
    # Line 5: XBH, 1B, 2B, 3B, HR
    ws["E6"] = "XBH"
    ws["F6"] = profile["XBH"] 
    ws["A7"] = "1B"
    ws["B7"] = profile["1B"] 
    ws["C7"] = "2B"
    ws["D7"] = profile["2B"]
    ws["E7"] = "3B"
    ws["F7"] = profile["3B"]
    ws["G7"] = "HR"
    ws["H7"] = profile["HR"]
    # Line 6: Line 5 percentages
    if profile["Hits"]:
        ws["B8"] = "{:.2f}%".format(profile["1B"]/profile["Hits"]*100)
        ws["D8"] = "{:.2f}%".format(profile["2B"]/profile["Hits"]*100)
        ws["F8"] = "{:.2f}%".format(profile["3B"]/profile["Hits"]*100)
        ws["H8"] = "{:.2f}%".format(profile["HR"]/profile["Hits"]*100)
    # Line 7: SO, L, S
    ws["A9"] = "SO"
    ws["B9"] = profile["K"]
    ws["C9"] = "Looking"
    ws["D9"] = profile["KL"]
    ws["E9"] = "Swinging"
    ws["F9"] = profile["KS"]
    # Line 8: Line 7 percentages
    if profile["PA"]:
        ws["B10"] = "{:.2f}%".format(profile["K"]/profile["PA"]*100)
        ws["D10"] = "{:.2f}%".format(profile["KL"]/profile["PA"]*100)
        ws["F10"] = "{:.2f}%".format(profile["KS"]/profile["PA"]*100)
    # Line 9: ROE, SF, FC
    ws["A11"] = "ROE"
    ws["B11"] = profile["ROE"]
    ws["C11"] = "SF"
    ws["D11"] = profile["SF"]
    ws["E11"] = "FC"
    ws["F11"] = profile["FC"]
    # IFH
    ws["A12"] = "IFH"
    ws["B12"] = profile["IFH"]
    ws["C12"] = "IFHO"
    ws["D12"] = profile["IFHO"]
    ws["E12"] = "IFH%"
    ws["F12"] = "{:.3f}%".format(profile["IFHP"])
    # Bunts
    ws["A13"] = "Steal Attempts"
    ws["B13"] = profile["CS"] + profile["SB"]
    ws["C13"] = "SB"
    ws["D13"] = profile["SB"]
    ws["E13"] = "CS"
    ws["F13"] = profile["CS"]
    ws["G13"] = "Steal %"
    if profile["CS"] + profile["SB"]:
        ws["H13"] = "{:.2f}%".format(profile["SB"]/(profile["SB"]+profile["CS"])*100)
    else:
        ws["H13"] = "N/A"
    ws["C14"] = "Steal 2"
    ws["D14"] = profile["SB2"]
    ws["E14"] = "Steal 3"
    ws["F14"] = profile["SB3"]
    ws["G14"] = "Steal 4"
    ws["H14"] = profile["SB4"]
    # Line 10: Bunts, Safe, Outs, Errors, Sacrifice
    ws["A15"] = "Safe Bunt"
    ws["B15"] = profile["Bunt"]["Safe"]
    ws["C15"] = "Out Bunt"
    ws["D15"] = profile["Bunt"]["Out"]
    ws["E15"] = "Errors"
    ws["F15"] = profile["Bunt"]["Error"]
    ws["G15"] = "Sacrifice"
    ws["H15"] = profile["Bunt"]["SAC"]
    # Line 11: Line 10 percentages
    if profile["Type"]["Bunt"]:
        ws["B16"] = "{:.2f}%".format(profile["Bunt"]["Safe"]/profile["Type"]["Bunt"]*100)
        ws["D16"] = "{:.2f}%".format(profile["Bunt"]["Out"]/profile["Type"]["Bunt"]*100)
        ws["F16"] = "{:.2f}%".format(profile["Bunt"]["Error"]/profile["Type"]["Bunt"]*100)
        ws["H16"] = "{:.2f}%".format(profile["Bunt"]["SAC"]/profile["Type"]["Bunt"]*100)
    # Line 12: Hit Types
    ws["A17"] = "Hit Types"
    ws["B17"] = "GB"
    ws["C17"] = "FB"
    ws["D17"] = "LD"
    ws["E17"] = "PF"
    ws["F17"] = "Bunts"
    # Line 13: Line 12 numbers
    ws["B18"] = profile["Type"]["GB"]
    ws["C18"] = profile["Type"]["FB"]
    ws["D18"] = profile["Type"]["LD"]
    ws["E18"] = profile["Type"]["PF"]
    ws["F18"] = profile["Type"]["Bunt"]
    # Line 14: Line 12 percentages
    total = (profile["Type"]["GB"] + profile["Type"]["GB"] + profile["Type"]["FB"] + 
        profile["Type"]["LD"] + profile["Type"]["PF"] + profile["Type"]["Bunt"])
    if total:
        ws["B19"] = "{:.2f}%".format(profile["Type"]["GB"]/total*100)
        ws["C19"] = "{:.2f}%".format(profile["Type"]["FB"]/total*100)
        ws["D19"] = "{:.2f}%".format(profile["Type"]["LD"]/total*100)
        ws["E19"] = "{:.2f}%".format(profile["Type"]["PF"]/total*100)
        ws["F19"] = "{:.2f}%".format(profile["Type"]["Bunt"]/total*100)
    # Line 15: Locations
    ws["A20"] = "Location:"
    ws["B20"] = "1"
    ws["C20"] = "2"
    ws["D20"] = "3"
    ws["E20"] = "4"
    ws["F20"] = "5"
    ws["G20"] = "6"
    ws["B23"] = "7"
    ws["C23"] = "8"
    ws["D23"] = "9"
    ws["E23"] = "78"
    ws["F23"] = "89"
    # Line 16: Line 15 numbers
    ws["B21"] = profile["Location"]["1"]
    ws["C21"] = profile["Location"]["2"]
    ws["D21"] = profile["Location"]["3"]
    ws["E21"] = profile["Location"]["4"]
    ws["F21"] = profile["Location"]["5"]
    ws["G21"] = profile["Location"]["6"]
    ws["B24"] = profile["Location"]["7"]
    ws["C24"] = profile["Location"]["8"]
    ws["D24"] = profile["Location"]["9"]
    ws["E24"] = profile["Location"]["78"]
    ws["F24"] = profile["Location"]["89"]
    # Line 17: Line 15 percentages
    total = (profile["Location"]["1"] + profile["Location"]["2"] + profile["Location"]["3"] + 
        profile["Location"]["4"] + profile["Location"]["5"] + profile["Location"]["6"] +
        profile["Location"]["7"] + profile["Location"]["8"] + profile["Location"]["9"] +
        profile["Location"]["78"] + profile["Location"]["89"])
    if total:
        ws["B22"] = "{:.2f}%".format(profile["Location"]["1"]/total*100)
        ws["C22"] = "{:.2f}%".format(profile["Location"]["2"]/total*100)
        ws["D22"] = "{:.2f}%".format(profile["Location"]["3"]/total*100)
        ws["E22"] = "{:.2f}%".format(profile["Location"]["4"]/total*100)
        ws["F22"] = "{:.2f}%".format(profile["Location"]["5"]/total*100)
        ws["G22"] = "{:.2f}%".format(profile["Location"]["6"]/total*100)
        ws["B25"] = "{:.2f}%".format(profile["Location"]["7"]/total*100)
        ws["C25"] = "{:.2f}%".format(profile["Location"]["8"]/total*100)
        ws["D25"] = "{:.2f}%".format(profile["Location"]["9"]/total*100)
        ws["E25"] = "{:.2f}%".format(profile["Location"]["78"]/total*100)
        ws["F25"] = "{:.2f}%".format(profile["Location"]["89"]/total*100)
    
    # Tab Colors
    wsprops = ws.sheet_properties
    if profile["PA"] == 0:
        wsprops.tabColor = color_3  
    elif profile["PA"] > 0 and profile["PA"] < 100:
        wsprops.tabColor = color_2
    elif profile["PA"] >= 100:
        wsprops.tabColor = color_1

    # Spray Chart
    img = Image("/home/ilan/Documents/pbp/images/spray_chart_base_2.png")
    img.width = 275
    img.height = 400
    ws.add_image(img, "B27")

    total = (profile["Location"]["1"] + profile["Location"]["2"] + profile["Location"]["3"] + 
        profile["Location"]["4"] + profile["Location"]["5"] + profile["Location"]["6"] +
        profile["Location"]["7"] + profile["Location"]["8"] + profile["Location"]["9"] +
        profile["Location"]["78"] + profile["Location"]["89"])
    if total:
        location_1 = "C40"
        location_2 = "D44"
        location_3 = "E40"
        location_4 = "D37"
        location_5 = "C37"
        location_6 = "B40"
        location_7 = "B31"
        location_8 = "C30"
        location_9 = "D31"

        ws[location_1] = "{:.2f}%".format(profile["Location"]["1"]/total*100)
        ws[location_2] = "{:.2f}%".format(profile["Location"]["2"]/total*100)
        ws[location_3] = "{:.2f}%".format(profile["Location"]["3"]/total*100)
        ws[location_4] = "{:.2f}%".format(profile["Location"]["4"]/total*100)
        ws[location_5] = "{:.2f}%".format(profile["Location"]["5"]/total*100)
        ws[location_6] = "{:.2f}%".format(profile["Location"]["6"]/total*100)
        ws[location_7] = "{:.2f}%".format(profile["Location"]["7"]/total*100)
        ws[location_8] = "{:.2f}%".format(profile["Location"]["8"]/total*100)
        ws[location_9] = "{:.2f}%".format(profile["Location"]["9"]/total*100)

    return True

def create_team_leaders_page(ws, profiles):
    ws["A1"] = "Team Leaders"
    ws["A3"] = "AVG"
    (ws["B3"], ws["C3"]) = find_max(profiles, "AVG")
    ws["A4"] = "OBP"
    (ws["B4"], ws["C4"]) = find_max(profiles, "OBP")
    ws["A5"] = "SLG"
    (ws["B5"], ws["C5"]) = find_max(profiles, "SLG")
    ws["A6"] = "OPS"
    (ws["B6"], ws["C6"]) = find_max(profiles, "OPS")
    ws["A7"] = "BABIP"
    (ws["B7"], ws["C7"]) = find_max(profiles, "BABIP")
    ws["A8"] = "PA"
    (ws["B8"], ws["C8"]) = find_max(profiles, "PA")
    ws["A9"] = "AB"
    (ws["B9"], ws["C9"]) = find_max(profiles, "AB")
    ws["A10"] = "Hits"
    (ws["B10"], ws["C10"]) = find_max(profiles, "Hits")
    ws["A11"] = "XBH"
    (ws["B11"], ws["C11"]) = find_max(profiles, "XBH")
    ws["A12"] = "1B"
    (ws["B12"], ws["C12"]) = find_max(profiles, "1B")
    ws["A13"] = "2B"
    (ws["B13"], ws["C13"]) = find_max(profiles, "2B")
    ws["A14"] = "3B"
    (ws["B14"], ws["C14"]) = find_max(profiles, "3B")
    ws["A15"] = "HR"
    (ws["B15"], ws["C15"]) = find_max(profiles, "HR")
    ws["A16"] = "BB"
    (ws["B16"], ws["C16"]) = find_max(profiles, "BB")
    ws["A17"] = "HBP"
    (ws["B17"], ws["C17"]) = find_max(profiles, "HBP")
    ws["A18"] = "K"
    (ws["B18"], ws["C18"]) = find_max(profiles, "K")
    ws["A19"] = "KL"
    (ws["B19"], ws["C19"]) = find_max(profiles, "KL")
    ws["A20"] = "KS"
    (ws["B20"], ws["C20"]) = find_max(profiles, "KS")
    ws["A21"] = "ROE"
    (ws["B21"], ws["C21"]) = find_max(profiles, "ROE")    
    ws["A22"] = "SF"
    (ws["B22"], ws["C22"]) = find_max(profiles, "SF")    
    ws["A23"] = "FC"
    (ws["B23"], ws["C23"]) = find_max(profiles, "FC")        
    ws["A24"] = "IFH"
    (ws["B24"], ws["C24"]) = find_max(profiles, "IFH")
    ws["A25"] = "IFH%"
    (ws["B25"], ws["C25"]) = find_max(profiles, "IFHP")
    ws["A26"] = "SB"
    (ws["B26"], ws["C26"]) = find_max(profiles, "SB")    
    ws["A27"] = "CS"
    (ws["B27"], ws["C27"]) = find_max(profiles, "CS")    
    ws["A28"] = "Stealing 2nd"
    (ws["B28"], ws["C28"]) = find_max(profiles, "SB2")    
    ws["A29"] = "Stealing 3rd"
    (ws["B29"], ws["C29"]) = find_max(profiles, "SB3")    
    ws["A30"] = "Stealing Home"
    (ws["B30"], ws["C30"]) = find_max(profiles, "SB4")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    return True

def create_legal_page(ws):
    ws["A1"] = disclaimer_1
    ws["A2"] = disclaimer_2

    return True

def find_max(profiles, stat):
    max = 0
    name = ""
    for profile in profiles:
        if profile[stat] >= max:
            max = profile[stat]
            name = profile["f_name"] + " " + profile["name"]
    if max == 0:
        return ("N/A", 0)
    return (name, max)